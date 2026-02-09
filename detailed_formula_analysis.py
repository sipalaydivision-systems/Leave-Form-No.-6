import openpyxl
from pathlib import Path

def detailed_formula_analysis(filepath):
    """Detailed analysis of Excel leave card formulas"""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)
        ws = wb.active
        
        print(f"\n{'='*90}")
        print(f"FILE: {Path(filepath).name}")
        print(f"{'='*90}")
        print(f"Sheet Name: {ws.title}")
        print(f"Dimensions: {ws.dimensions}")
        
        print(f"\n{'STRUCTURE OVERVIEW':^90}")
        print(f"{'-'*90}")
        
        # Show first 20 rows to understand structure
        print("Row Layout:")
        for row_idx in range(1, 20):
            row_data = []
            for col_idx in range(1, 11):  # Columns A-J
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    val = str(cell.value)[:30]
                    row_data.append(f"{cell.coordinate}:{val}")
            if row_data:
                print(f"  Row {row_idx}: {', '.join(row_data)}")
        
        print(f"\n{'FORMULAS ANALYSIS':^90}")
        print(f"{'-'*90}")
        
        formulas = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    formulas.append({
                        'coordinate': cell.coordinate,
                        'formula': cell.value,
                        'row': cell.row
                    })
        
        if formulas:
            print(f"Total formulas: {len(formulas)}\n")
            
            # Group by type of formula
            unique_formulas = {}
            for f in formulas:
                formula = f['formula']
                if formula not in unique_formulas:
                    unique_formulas[formula] = []
                unique_formulas[formula].append(f['coordinate'])
            
            print("Unique Formula Patterns:")
            for formula, cells in sorted(unique_formulas.items()):
                print(f"\n  Formula: {formula}")
                print(f"  Used in cells: {', '.join(cells[:5])}" + (f" ... and {len(cells)-5} more" if len(cells) > 5 else ""))
        else:
            print("No formulas found (values only)")
        
        return formulas
        
    except Exception as e:
        print(f"ERROR: {e}")
        return []

# Analyze both files
non_teaching_file = r'e:\Division Files\Leave Form No. 6\OneDrive_2026-02-05\LEAVE CARD-NON-TEACHING PERSONNEL\ACUHIDO, ELIZA B.xlsx'
teaching_file = r'e:\Division Files\Leave Form No. 6\OneDrive_2026-02-05_1\LEAVE CARD-TEACHING PERSONNEL\ABANILLA, FE.xlsx'

print("\n" + "="*90)
print("DETAILED LEAVE CARD FORMULA ANALYSIS")
print("="*90)

formulas1 = detailed_formula_analysis(non_teaching_file)
formulas2 = detailed_formula_analysis(teaching_file)

# Summary explanation
print("\n" + "="*90)
print("FORMULA LOGIC EXPLANATION")
print("="*90)
print("""
The Leave Card formulas work as follows:

Column H (VACATION BALANCE):
  - H14 = B14  [Initial vacation earned for the period]
  - H15 onwards: H = Previous Balance - Forced Leave - Vacation Spent + New Vacation Earned
    Formula: H15 = H14 - F15 - D15 + B15

Column I (SICK LEAVE BALANCE):
  - I14 = C14  [Initial sick leave earned for the period]
  - I15 onwards: I = Previous Balance - Sick Leave Spent + New Sick Leave Earned
    Formula: I15 = I14 - E15 + C15

Column J (TOTAL BALANCE):
  - J = H + I  [Sum of vacation and sick leave balances]
    Formula: J14 = H14 + I14, then J15 = H15 + I15, etc.

Key Columns:
  - B: Vacation Leave Earned
  - C: Sick Leave Earned
  - D: Vacation Leave Spent
  - E: Sick Leave Spent
  - F: Forced Leave Used
  - H: Running Balance of Vacation Leave
  - I: Running Balance of Sick Leave
  - J: Total Leave Balance

This creates a running balance system that:
1. Starts with initial balances (H14, I14)
2. Tracks new leaves earned (B, C columns)
3. Subtracts leaves taken (D, E, F columns)
4. Maintains cumulative balances for each leave type
5. Provides total available balance
""")
