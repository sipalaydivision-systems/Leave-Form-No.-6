import openpyxl
import json
import os
from datetime import datetime

os.chdir(r'f:\Division Files\Leave Form No. 6')

# Read Excel file
wb = openpyxl.load_workbook('SDO-SIPALAY-MASTER-FILE-as-of-January-2026.xlsx')
sheet = wb.active

print(f"Reading from sheet: {sheet.title}")
print(f"Total rows with data: {sheet.max_row}")

employees = []
employee_id = 1

# Skip header row and read data
for row_idx in range(2, sheet.max_row + 1):
    try:
        employee_number = sheet[f'A{row_idx}'].value
        last_name = sheet[f'B{row_idx}'].value
        first_name = sheet[f'C{row_idx}'].value
        middle_name = sheet[f'D{row_idx}'].value
        position = sheet[f'G{row_idx}'].value
        salary_grade = sheet[f'H{row_idx}'].value
        step = sheet[f'I{row_idx}'].value
        salary = sheet[f'J{row_idx}'].value
        school_id = sheet[f'K{row_idx}'].value
        school_name = sheet[f'L{row_idx}'].value
        
        # Skip if critical data missing
        if not last_name or not first_name or not position:
            continue
            
        # Convert types
        if salary_grade:
            try:
                salary_grade = int(salary_grade)
            except:
                salary_grade = None
                
        if step:
            try:
                step = int(step)
            except:
                step = 1
        else:
            step = 1
            
        if salary:
            try:
                salary = int(salary)
            except:
                salary = 0
                
        # Build full name
        full_name = f"{first_name} {middle_name if middle_name else ''} {last_name}".strip()
        
        # Determine office/district
        office = school_name if school_name else "SDO"
        district = "Sipalay"
        
        # Create employee record
        employee = {
            "id": employee_id,
            "officeCode": str(school_id) if school_id else "SDO",
            "office": office,
            "district": district,
            "lastName": last_name.strip() if last_name else "",
            "firstName": first_name.strip() if first_name else "",
            "middleName": middle_name.strip() if middle_name else "",
            "fullName": full_name,
            "position": position.strip() if position else "",
            "salaryGrade": salary_grade,
            "step": step,
            "salary": salary,
            "email": f"{first_name.lower()}.{last_name.lower()}@deped.gov.ph".replace(" ", ""),
            "createdAt": datetime.now().isoformat(),
            "leaveCredits": 10,
            "lastLeaveUpdate": datetime.now().isoformat()
        }
        
        employees.append(employee)
        employee_id += 1
        
        if employee_id % 100 == 0:
            print(f"  Processed {employee_id - 1} employees...")
            
    except Exception as e:
        print(f"  Error on row {row_idx}: {e}")
        continue

print(f"\nTotal employees extracted: {len(employees)}")

# Write to employees.json
with open('employees.json', 'w') as f:
    json.dump(employees, f, indent=2)

print("Wrote to employees.json")

# Print summary
print("\nSample records:")
for emp in employees[:5]:
    print(f"  {emp['fullName']} - {emp['position']} (SG {emp['salaryGrade']} Step {emp['step']} - ₱{emp['salary']:,})")

# Print salary grid summary
print("\nSalary Grade Summary:")
salary_grades = {}
for emp in employees:
    if emp['salaryGrade']:
        if emp['salaryGrade'] not in salary_grades:
            salary_grades[emp['salaryGrade']] = {}
        key = f"Step {emp['step']}"
        if key not in salary_grades[emp['salaryGrade']]:
            salary_grades[emp['salaryGrade']][key] = emp['salary']

for sg in sorted(salary_grades.keys()):
    print(f"  SG {sg}: {salary_grades[sg]}")

wb.close()
