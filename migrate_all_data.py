"""
Comprehensive Migration Script
===============================
Reads ALL Excel leave card files (non-teaching + teaching personnel),
extracts leave card transactions and CTO records,
creates temporary employee portal accounts,
and generates a credentials Excel file for distribution.

Usage: python migrate_all_data.py
"""

import openpyxl
import json
import os
import hashlib
import secrets
import string
import re
from datetime import datetime
from copy import deepcopy

# ============================================================
# CONFIGURATION
# ============================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')

NON_TEACHING_DIR = os.path.join(BASE_DIR, 'OneDrive_2026-02-05', 'LEAVE CARD-NON-TEACHING PERSONNEL')
TEACHING_DIR = os.path.join(BASE_DIR, 'OneDrive_2026-02-05_1', 'LEAVE CARD-TEACHING PERSONNEL')

EMPLOYEES_FILE = os.path.join(DATA_DIR, 'employees.json')
USERS_FILE = os.path.join(DATA_DIR, 'users.json')
LEAVECARDS_FILE = os.path.join(DATA_DIR, 'leavecards.json')
CTO_RECORDS_FILE = os.path.join(DATA_DIR, 'cto-records.json')
SO_RECORDS_FILE = os.path.join(DATA_DIR, 'so-records.json')

CREDENTIALS_OUTPUT = os.path.join(BASE_DIR, 'EMPLOYEE_TEMP_CREDENTIALS.xlsx')

# ============================================================
# UTILITY FUNCTIONS
# ============================================================

def read_json(filepath):
    """Read a JSON file and return its content."""
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
            if content.startswith('\ufeff'):
                content = content[1:]
            return json.loads(content)
    except Exception as e:
        print(f"  [WARN] Error reading {filepath}: {e}")
        return []

def write_json(filepath, data):
    """Write data to a JSON file."""
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

def hash_password_with_salt(password):
    """Hash a password with a random salt (matches server.js hashPasswordWithSalt)."""
    salt = secrets.token_hex(16)
    hash_val = hashlib.sha256((salt + password).encode()).hexdigest()
    return f"{salt}:{hash_val}"

def generate_temp_password(length=8):
    """Generate a temporary password that's easy to communicate."""
    # Format: 3 uppercase + 3 digits + 2 lowercase for readability
    upper = ''.join(secrets.choice(string.ascii_uppercase) for _ in range(3))
    digits = ''.join(secrets.choice(string.digits) for _ in range(3))
    lower = ''.join(secrets.choice(string.ascii_lowercase) for _ in range(2))
    pwd = upper + digits + lower
    # Shuffle it
    pwd_list = list(pwd)
    secrets.SystemRandom().shuffle(pwd_list)
    return ''.join(pwd_list)

def normalize_name(filename):
    """Extract clean name from filename like 'ACUHIDO, ELIZA B..xlsx' -> 'ACUHIDO, ELIZA B.'"""
    name = os.path.splitext(filename)[0]
    # Remove common suffixes
    for suffix in ['-SO', '- SO', '-ASDS', '-OIC ASDS', '-PROP. CUST.', '-Prop.Cust.', 
                   '-PROP.CUST.', '-CANTURAY ES', '-CANTACA ES', '-BINULIG ES',
                   '-CALANGCANG ANNEX', '-S.O.', ' - S.O.', ' - SO.', '-SO.']:
        if name.upper().endswith(suffix.upper()):
            name = name[:len(name)-len(suffix)]
            break
    return name.strip()

def name_to_lastname(name_str):
    """Extract last name from 'LASTNAME, FIRSTNAME' format."""
    if ',' in name_str:
        return name_str.split(',')[0].strip().upper()
    return name_str.strip().upper()

def name_to_firstname(name_str):
    """Extract first name from 'LASTNAME, FIRSTNAME' format."""
    if ',' in name_str:
        parts = name_str.split(',', 1)
        return parts[1].strip().upper() if len(parts) > 1 else ''
    return ''

def safe_float(val, default=0):
    """Safely convert a value to float."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    try:
        # Handle strings like "3 DAYS & 3 HRS"
        s = str(val).strip()
        if not s:
            return default
        # Try to extract just the number
        match = re.match(r'^[\d.]+', s)
        if match:
            return float(match.group())
        return default
    except:
        return default

def safe_str(val):
    """Safely convert a cell value to string."""
    if val is None:
        return ''
    return str(val).strip()


# ============================================================
# EXCEL PARSING: LEAVE CARD
# ============================================================

def find_leave_card_sheet(wb):
    """Find the best leave card sheet in a workbook."""
    # Priority order for leave card sheet names
    priority_names = [
        'CORRECTED', 'CATIGAN', 'DUSARAN', 'VALIENTE',
        'Leave Card CORRECTED', 'Leave Card CORRECTED ', 'LEAVE CARD CORRECTED',
        'Leave Card (NE) CORRECTED', 'Leave Card FINAL', 'FOR RETIREMENT (2)',
        'LEAVE CARD', 'Leave Card (NE)', 'Leave Card (NE) (2)', 'Leave Card (NE) (3)',
        'Leave Card', 'Leave Card ', 'Leave Card as TIC',
    ]
    
    sheets = wb.sheetnames
    
    # First, check for CORRECTED versions (most accurate)
    for name in sheets:
        n = name.strip().upper()
        if 'CORRECT' in n and ('LEAVE' in n or n == 'CORRECTED'):
            return name
    
    # Check for employee-name based sheets (custom corrected)
    for name in sheets:
        n = name.strip().upper()
        if n in ['CATIGAN', 'DUSARAN', 'VALIENTE']:
            return name
    
    # Check standard names
    for pname in priority_names:
        for name in sheets:
            if name.strip().lower() == pname.strip().lower():
                return name
    
    # Fallback: any sheet with "leave card" in name
    for name in sheets:
        n = name.strip().upper()
        if 'LEAVE CARD' in n and 'BLANK' not in n and 'OLD' not in n and 'WRONG' not in n:
            return name
    
    # Last resort: Sheet1 might be a leave card
    if 'Sheet1' in sheets:
        return 'Sheet1'
    
    return sheets[0] if sheets else None

def find_cto_sheet(wb):
    """Find the CTO sheet in a workbook."""
    sheets = wb.sheetnames
    
    # Check for latest/corrected CTO first
    for name in sheets:
        n = name.strip().upper()
        if 'CTO' in n and ('LATEST' in n or 'CORRECT' in n):
            return name
    
    # Then standard CTO variations
    for name in sheets:
        n = name.strip().upper()
        if n.strip() in ['CTO', 'CTO (2)', 'CTO 2024-2025', 'COMPENSATORY TIME OFF (CTO)']:
            return name
    
    for name in sheets:
        n = name.strip().upper()
        if 'CTO' in n and 'WRONG' not in n:
            return name
    
    return None

def is_leave_card_sheet(ws):
    """Check if a worksheet appears to be a leave card (not CTO/VSC/Election)."""
    # Look for the header pattern of a leave card
    for row in range(1, min(20, ws.max_row + 1)):
        for col in range(1, min(12, ws.max_column + 1)):
            cell_val = safe_str(ws.cell(row=row, column=col).value).upper()
            if 'LEAVE EARNED' in cell_val or 'SERVICE LEAVE CARD' in cell_val:
                return True
            if 'VACATION' in cell_val and 'SICK' in cell_val:
                return True
    return False

def is_cto_vsc_sheet(ws):
    """Check if a worksheet is a CTO/VSC type sheet."""
    for row in range(1, min(20, ws.max_row + 1)):
        for col in range(1, min(12, ws.max_column + 1)):
            cell_val = safe_str(ws.cell(row=row, column=col).value).upper()
            if 'SPECIAL ORDER' in cell_val or 'NO. OF DAYS GRANTED' in cell_val:
                return True
            if 'COMPENSATORY TIME OFF' in cell_val or 'VSC' == cell_val.strip():
                return True
    return False

def parse_leave_card_sheet(ws, filename):
    """Parse a leave card worksheet and extract transactions."""
    transactions = []
    last_vl_balance = 0
    last_sl_balance = 0
    date_of_appointment = None
    
    # Find header row (look for VACATION/SICK headers)
    header_row = None
    data_start_row = None
    
    for row_num in range(1, min(20, ws.max_row + 1)):
        a_val = safe_str(ws.cell(row=row_num, column=1).value).upper()
        b_val = safe_str(ws.cell(row=row_num, column=2).value).upper()
        
        if 'DATE OF ORIGINAL APPOINTMENT' in a_val:
            # Extract date
            parts = a_val.split(':', 1)
            if len(parts) > 1:
                date_of_appointment = parts[1].strip()
        
        if 'PERIOD COVERED' in a_val:
            header_row = row_num
        
        if 'VACATION' in b_val and header_row:
            data_start_row = row_num + 1
            break
    
    if not data_start_row:
        # Try finding first ADD/LESS row
        for row_num in range(1, ws.max_row + 1):
            a_val = safe_str(ws.cell(row=row_num, column=1).value).upper()
            if a_val.startswith('ADD:') or a_val.startswith('LESS:'):
                data_start_row = row_num
                break
    
    if not data_start_row:
        return [], 0, 0, date_of_appointment
    
    # Parse data rows
    for row_num in range(data_start_row, ws.max_row + 1):
        a_val = safe_str(ws.cell(row=row_num, column=1).value)
        if not a_val:
            continue
        
        a_upper = a_val.upper().strip()
        
        # Skip non-data rows
        if a_upper.startswith('CERTIFIED') or a_upper.startswith('NOTE:') or \
           a_upper.startswith('MA.') or a_upper.startswith('ADMINISTRATIVE') or \
           a_upper.startswith('HRMO') or a_upper.startswith('DATE OF'):
            continue
        
        if not (a_upper.startswith('ADD:') or a_upper.startswith('LESS:')):
            # Could be a continuation or date of appointment
            if 'DATE OF ORIGINAL APPOINTMENT' in a_upper:
                parts = a_upper.split(':', 1)
                if len(parts) > 1:
                    date_of_appointment = parts[1].strip()
            continue
        
        txn_type = 'ADD' if a_upper.startswith('ADD:') else 'LESS'
        period = a_val.split(':', 1)[1].strip() if ':' in a_val else a_val
        
        # Read values from columns B through J
        vl_earned = safe_float(ws.cell(row=row_num, column=2).value)   # B - VL Earned
        sl_earned = safe_float(ws.cell(row=row_num, column=3).value)   # C - SL Earned
        vl_spent = safe_float(ws.cell(row=row_num, column=4).value)    # D - VL Spent
        sl_spent = safe_float(ws.cell(row=row_num, column=5).value)    # E - SL Spent
        forced = safe_float(ws.cell(row=row_num, column=6).value)      # F - Forced Leave
        spl = safe_float(ws.cell(row=row_num, column=7).value)         # G - Special Privilege
        vl_balance = safe_float(ws.cell(row=row_num, column=8).value)  # H - VL Balance
        sl_balance = safe_float(ws.cell(row=row_num, column=9).value)  # I - SL Balance
        total = safe_float(ws.cell(row=row_num, column=10).value)      # J - Total
        
        txn = {
            'type': txn_type,
            'periodCovered': period,
            'vlEarned': vl_earned,
            'slEarned': sl_earned,
            'vlSpent': vl_spent,
            'slSpent': sl_spent,
            'forcedLeave': forced,
            'splUsed': spl,
            'vlBalance': vl_balance,
            'slBalance': sl_balance,
            'total': total
        }
        transactions.append(txn)
        
        # Track latest balances
        if vl_balance > 0 or sl_balance > 0 or total > 0:
            last_vl_balance = vl_balance
            last_sl_balance = sl_balance
    
    return transactions, last_vl_balance, last_sl_balance, date_of_appointment

def parse_cto_sheet(ws, filename):
    """Parse a CTO/VSC worksheet and extract CTO records."""
    records = []
    
    # Find data start
    data_start_row = None
    for row_num in range(1, min(20, ws.max_row + 1)):
        a_val = safe_str(ws.cell(row=row_num, column=1).value).upper()
        if a_val.startswith('ADD:') or a_val.startswith('LESS:'):
            data_start_row = row_num
            break
    
    if not data_start_row:
        return records
    
    # CTO format: A=period, B=SO details, F=days granted, H=days used, J=balance, L=total
    current_so = None
    
    for row_num in range(data_start_row, ws.max_row + 1):
        a_val = safe_str(ws.cell(row=row_num, column=1).value)
        if not a_val:
            continue
        
        a_upper = a_val.upper().strip()
        
        if a_upper.startswith('NOTE:') or a_upper.startswith('CERTIFIED') or \
           a_upper.startswith('MA.') or a_upper.startswith('ADMINISTRATIVE') or \
           a_upper.startswith('HRMO'):
            continue
        
        if not (a_upper.startswith('ADD:') or a_upper.startswith('LESS:')):
            continue
        
        txn_type = 'ADD' if a_upper.startswith('ADD:') else 'LESS'
        period = a_val.split(':', 1)[1].strip() if ':' in a_val else a_val
        
        so_details = safe_str(ws.cell(row=row_num, column=2).value)
        days_granted = safe_float(ws.cell(row=row_num, column=6).value)
        days_used = safe_float(ws.cell(row=row_num, column=8).value)
        balance = safe_float(ws.cell(row=row_num, column=10).value)
        total = safe_float(ws.cell(row=row_num, column=12).value)
        
        record = {
            'type': txn_type,
            'periodCovered': period,
            'soDetails': so_details,
            'daysGranted': days_granted,
            'daysUsed': days_used,
            'balance': balance,
            'total': total
        }
        records.append(record)
    
    return records


# ============================================================
# EMPLOYEE MATCHING
# ============================================================

def match_employee(filename, employees):
    """Match an Excel filename to an employee record."""
    clean_name = normalize_name(filename)
    last_name = name_to_lastname(clean_name)
    first_name = name_to_firstname(clean_name)
    
    # Try exact last name + first name match
    candidates = []
    for emp in employees:
        emp_lastname = emp.get('lastName', '').upper().strip()
        emp_firstname = emp.get('firstName', '').upper().strip()
        emp_fullname = emp.get('fullName', '').upper().strip()
        
        if emp_lastname == last_name:
            # Check if first name matches
            if first_name:
                fn_upper = first_name.upper()
                # Exact match
                if emp_firstname == fn_upper:
                    return emp
                # Partial match (first name starts with or contains)
                if emp_firstname.startswith(fn_upper.split()[0]) or fn_upper.split()[0] in emp_firstname:
                    candidates.append(emp)
                # Check full name contains
                elif fn_upper.split()[0] in emp_fullname:
                    candidates.append(emp)
            else:
                candidates.append(emp)
    
    if len(candidates) == 1:
        return candidates[0]
    elif len(candidates) > 1:
        # Try to narrow down further with first name
        if first_name:
            for c in candidates:
                if first_name.upper().split()[0] == c.get('firstName', '').upper().split()[0]:
                    return c
        return candidates[0]  # Return first match
    
    return None


# ============================================================
# MAIN MIGRATION
# ============================================================

def main():
    print("=" * 70)
    print("LEAVE CARD DATA MIGRATION TOOL")
    print("=" * 70)
    print(f"Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    # Load existing data
    employees = read_json(EMPLOYEES_FILE)
    existing_users = read_json(USERS_FILE)
    existing_leavecards = read_json(LEAVECARDS_FILE)
    existing_cto = read_json(CTO_RECORDS_FILE)
    existing_so = read_json(SO_RECORDS_FILE)
    
    print(f"Existing employees: {len(employees)}")
    print(f"Existing user accounts: {len(existing_users)}")
    print(f"Existing leave cards: {len(existing_leavecards)}")
    print(f"Existing CTO records: {len(existing_cto)}")
    print()
    
    # Collect all Excel files
    excel_files = []
    
    # Non-teaching
    if os.path.exists(NON_TEACHING_DIR):
        for f in sorted(os.listdir(NON_TEACHING_DIR)):
            if f.endswith('.xlsx') and not f.startswith('~$'):
                excel_files.append({
                    'path': os.path.join(NON_TEACHING_DIR, f),
                    'filename': f,
                    'type': 'non-teaching'
                })
    
    # Teaching
    if os.path.exists(TEACHING_DIR):
        for f in sorted(os.listdir(TEACHING_DIR)):
            if f.endswith('.xlsx') and not f.startswith('~$'):
                excel_files.append({
                    'path': os.path.join(TEACHING_DIR, f),
                    'filename': f,
                    'type': 'teaching'
                })
    
    print(f"Found {len(excel_files)} Excel files to process")
    print()
    
    # Track results
    new_leavecards = []
    new_cto_records = []
    temp_credentials = []
    new_users = list(existing_users)
    existing_emails = {u['email'].lower() for u in existing_users}
    existing_leavecard_emails = {lc.get('email', '').lower() for lc in existing_leavecards}
    
    matched_count = 0
    unmatched_files = []
    processed = 0
    errors = []
    
    id_counter = int(datetime.now().timestamp() * 1000)
    
    for i, ef in enumerate(excel_files):
        filepath = ef['path']
        filename = ef['filename']
        personnel_type = ef['type']
        
        processed += 1
        if processed % 50 == 0:
            print(f"  Processing {processed}/{len(excel_files)}...")
        
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            errors.append(f"Cannot open {filename}: {e}")
            continue
        
        # Match to employee
        emp = match_employee(filename, employees)
        
        if not emp:
            unmatched_files.append(filename)
            wb.close()
            continue
        
        matched_count += 1
        emp_email = emp.get('email', '').lower()
        emp_id = emp.get('id')
        emp_name = emp.get('fullName', normalize_name(filename))
        emp_position = emp.get('position', '')
        emp_office = emp.get('office', '')
        emp_salary = emp.get('salary', 0)
        emp_sg = emp.get('salaryGrade', 0)
        emp_step = emp.get('step', 1)
        
        # ---- PARSE LEAVE CARD ----
        leave_sheet_name = find_leave_card_sheet(wb)
        transactions = []
        vl_balance = 0
        sl_balance = 0
        date_of_appointment = None
        
        if leave_sheet_name:
            ws = wb[leave_sheet_name]
            # Verify it's actually a leave card (not CTO/VSC)
            if is_leave_card_sheet(ws):
                transactions, vl_balance, sl_balance, date_of_appointment = parse_leave_card_sheet(ws, filename)
            elif is_cto_vsc_sheet(ws):
                # This is actually a CTO/VSC sheet misidentified, parse as CTO
                pass
        
        # ---- PARSE CTO ----
        cto_sheet_name = find_cto_sheet(wb)
        cto_records = []
        
        if cto_sheet_name:
            ws_cto = wb[cto_sheet_name]
            cto_records = parse_cto_sheet(ws_cto, filename)
        
        # Also check for VSC/VLC sheets (same format as CTO)
        for sheet_name in wb.sheetnames:
            sn = sheet_name.strip().upper()
            if sn in ['VSC', 'VSC (2)', 'VLC', 'VACATION LEAVE CREDITS', 'CONVERTED VSC', 'VLC']:
                try:
                    ws_vsc = wb[sheet_name]
                    vsc_records = parse_cto_sheet(ws_vsc, filename)
                    cto_records.extend(vsc_records)
                except:
                    pass
        
        wb.close()
        
        # ---- BUILD LEAVE CARD RECORD ----
        if emp_email and emp_email not in existing_leavecard_emails:
            # Calculate totals from transactions
            total_vl_earned = sum(t['vlEarned'] for t in transactions if t['type'] == 'ADD')
            total_sl_earned = sum(t['slEarned'] for t in transactions if t['type'] == 'ADD')
            total_vl_spent = sum(t['vlSpent'] for t in transactions if t['type'] == 'LESS')
            total_sl_spent = sum(t['slSpent'] for t in transactions if t['type'] == 'LESS')
            total_forced = sum(t['forcedLeave'] for t in transactions if t['type'] == 'LESS')
            total_spl = sum(t['splUsed'] for t in transactions if t['type'] == 'LESS')
            
            leavecard = {
                'employeeId': emp_email,
                'email': emp_email,
                'employeeName': emp_name,
                'personnelType': personnel_type,
                'dateOfAppointment': date_of_appointment,
                'transactions': transactions,
                'vacationLeaveEarned': round(total_vl_earned, 3),
                'sickLeaveEarned': round(total_sl_earned, 3),
                'forceLeaveEarned': 5,
                'splEarned': 3,
                'vacationLeaveSpent': round(total_vl_spent + total_forced, 3),
                'sickLeaveSpent': round(total_sl_spent, 3),
                'forceLeaveSpent': 0,
                'splSpent': 0,
                'vl': round(vl_balance, 3),
                'sl': round(sl_balance, 3),
                'spl': 3,
                'others': 0,
                'mandatoryForced': 5,
                'forceLeaveYear': 2026,
                'splYear': 2026,
                'leaveUsageHistory': [],
                'migratedFromExcel': True,
                'sourceFile': filename,
                'createdAt': datetime.now().isoformat() + 'Z',
                'updatedAt': datetime.now().isoformat() + 'Z',
                'initialCreditsSource': 'excel-migration'
            }
            new_leavecards.append(leavecard)
            existing_leavecard_emails.add(emp_email)
        
        # ---- BUILD CTO RECORDS ----
        if cto_records and emp_email:
            for cto in cto_records:
                id_counter += 1
                cto_record = {
                    'id': id_counter,
                    'employeeId': emp_email,
                    'email': emp_email,
                    'type': cto['type'],
                    'soDetails': cto.get('soDetails', ''),
                    'soNumber': cto.get('soDetails', '')[:50] if cto.get('soDetails') else '',
                    'periodCovered': cto.get('periodCovered', ''),
                    'daysGranted': cto.get('daysGranted', 0),
                    'daysUsed': cto.get('daysUsed', 0),
                    'balance': cto.get('balance', 0),
                    'migratedFromExcel': True,
                    'sourceFile': filename,
                    'createdAt': datetime.now().isoformat() + 'Z'
                }
                new_cto_records.append(cto_record)
        
        # ---- CREATE TEMP USER ACCOUNT ----
        if emp_email and emp_email not in existing_emails:
            temp_password = generate_temp_password()
            id_counter += 1
            
            new_user = {
                'id': id_counter,
                'email': emp_email,
                'password': hash_password_with_salt(temp_password),
                'name': emp_name,
                'fullName': emp_name,
                'office': emp_office,
                'position': emp_position,
                'employeeNo': '',
                'salaryGrade': emp_sg,
                'step': emp_step,
                'salary': emp_salary,
                'role': 'user',
                'mustChangePassword': True,
                'createdAt': datetime.now().isoformat() + 'Z'
            }
            
            new_users.append(new_user)
            existing_emails.add(emp_email)
            
            temp_credentials.append({
                'name': emp_name,
                'email': emp_email,
                'tempPassword': temp_password,
                'position': emp_position,
                'office': emp_office,
                'personnelType': personnel_type
            })
    
    # ---- ALSO CREATE ACCOUNTS FOR EMPLOYEES NOT IN EXCEL ----
    print(f"\nChecking for employees without Excel files...")
    additional_creds = 0
    for emp in employees:
        emp_email = emp.get('email', '').lower()
        if emp_email and emp_email not in existing_emails:
            temp_password = generate_temp_password()
            id_counter += 1
            
            new_user = {
                'id': id_counter,
                'email': emp_email,
                'password': hash_password_with_salt(temp_password),
                'name': emp.get('fullName', ''),
                'fullName': emp.get('fullName', ''),
                'office': emp.get('office', ''),
                'position': emp.get('position', ''),
                'employeeNo': '',
                'salaryGrade': emp.get('salaryGrade', 0),
                'step': emp.get('step', 1),
                'salary': emp.get('salary', 0),
                'role': 'user',
                'mustChangePassword': True,
                'createdAt': datetime.now().isoformat() + 'Z'
            }
            
            new_users.append(new_user)
            existing_emails.add(emp_email)
            
            temp_credentials.append({
                'name': emp.get('fullName', ''),
                'email': emp_email,
                'tempPassword': temp_password,
                'position': emp.get('position', ''),
                'office': emp.get('office', ''),
                'personnelType': 'from-employees-db'
            })
            additional_creds += 1
    
    print(f"  Created {additional_creds} additional accounts for employees not in Excel")
    
    # ---- SAVE RESULTS ----
    print(f"\n{'=' * 70}")
    print("MIGRATION RESULTS")
    print(f"{'=' * 70}")
    print(f"Excel files processed: {processed}")
    print(f"Matched to employees: {matched_count}")
    print(f"Unmatched files: {len(unmatched_files)}")
    print(f"Errors: {len(errors)}")
    print(f"New leave cards created: {len(new_leavecards)}")
    print(f"New CTO records created: {len(new_cto_records)}")
    print(f"Total user accounts: {len(new_users)}")
    print(f"New temp credentials: {len(temp_credentials)}")
    print()
    
    # Write leave cards
    all_leavecards = existing_leavecards + new_leavecards
    write_json(LEAVECARDS_FILE, all_leavecards)
    print(f"✓ Saved {len(all_leavecards)} leave cards to leavecards.json")
    
    # Write CTO records
    all_cto = existing_cto + new_cto_records
    write_json(CTO_RECORDS_FILE, all_cto)
    print(f"✓ Saved {len(all_cto)} CTO records to cto-records.json")
    
    # Write users
    write_json(USERS_FILE, new_users)
    print(f"✓ Saved {len(new_users)} users to users.json")
    
    # ---- GENERATE CREDENTIALS EXCEL ----
    print(f"\nGenerating credentials Excel file...")
    
    cred_wb = openpyxl.Workbook()
    ws = cred_wb.active
    ws.title = "Temporary Credentials"
    
    # Header styling
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title row
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = "SCHOOLS DIVISION OF SIPALAY CITY - Leave Management System"
    title_cell.font = Font(bold=True, size=14, color="1F4E79")
    title_cell.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A2:G2')
    subtitle = ws['A2']
    subtitle.value = f"Temporary Login Credentials (Generated: {datetime.now().strftime('%m/%d/%Y %I:%M %p')})"
    subtitle.font = Font(italic=True, size=11, color="666666")
    subtitle.alignment = Alignment(horizontal='center')
    
    ws.merge_cells('A3:G3')
    note = ws['A3']
    note.value = "⚠ IMPORTANT: All employees must change their password upon first login. Portal URL: https://leave-management.up.railway.app"
    note.font = Font(bold=True, size=10, color="CC0000")
    note.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[3].height = 30
    
    # Headers
    headers = ['#', 'Full Name', 'Email', 'Temporary Password', 'Position', 'School/Office', 'Type']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Sort credentials by type, then name
    temp_credentials.sort(key=lambda x: (x.get('personnelType', ''), x.get('name', '')))
    
    # Data rows
    alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    
    for idx, cred in enumerate(temp_credentials, 1):
        row = idx + 5
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=cred['name']).border = thin_border
        ws.cell(row=row, column=3, value=cred['email']).border = thin_border
        
        pwd_cell = ws.cell(row=row, column=4, value=cred['tempPassword'])
        pwd_cell.border = thin_border
        pwd_cell.font = Font(name='Consolas', size=11, bold=True)
        
        ws.cell(row=row, column=5, value=cred['position']).border = thin_border
        ws.cell(row=row, column=6, value=cred['office']).border = thin_border
        ws.cell(row=row, column=7, value=cred['personnelType']).border = thin_border
        
        # Alternate row coloring
        if idx % 2 == 0:
            for col in range(1, 8):
                ws.cell(row=row, column=col).fill = alt_fill
    
    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 15
    
    # Summary row
    summary_row = len(temp_credentials) + 7
    ws.cell(row=summary_row, column=1, value=f"Total: {len(temp_credentials)} accounts")
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=11)
    
    cred_wb.save(CREDENTIALS_OUTPUT)
    print(f"✓ Saved credentials to {CREDENTIALS_OUTPUT}")
    
    # ---- PRINT UNMATCHED FILES ----
    if unmatched_files:
        print(f"\n{'=' * 70}")
        print(f"UNMATCHED FILES ({len(unmatched_files)}):")
        print(f"{'=' * 70}")
        for uf in unmatched_files[:30]:
            print(f"  - {uf}")
        if len(unmatched_files) > 30:
            print(f"  ... and {len(unmatched_files) - 30} more")
    
    if errors:
        print(f"\n{'=' * 70}")
        print(f"ERRORS ({len(errors)}):")
        print(f"{'=' * 70}")
        for err in errors[:20]:
            print(f"  - {err}")
    
    print(f"\n{'=' * 70}")
    print("MIGRATION COMPLETE!")
    print(f"{'=' * 70}")
    print(f"Finished at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"\nNext steps:")
    print(f"  1. Review EMPLOYEE_TEMP_CREDENTIALS.xlsx")
    print(f"  2. Distribute credentials to employees")
    print(f"  3. Push changes: git add -A && git commit && git push")
    print(f"  4. Employees login at https://leave-management.up.railway.app")
    print(f"  5. They must change password on first login")


if __name__ == '__main__':
    main()
