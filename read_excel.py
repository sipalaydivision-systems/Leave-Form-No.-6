import openpyxl
import json
import os

os.chdir(r'f:\Division Files\Leave Form No. 6')
wb = openpyxl.load_workbook('SDO-SIPALAY-MASTER-FILE-as-of-January-2026.xlsx')
print('Available sheets:', wb.sheetnames)

# Try to find the salary grid sheet
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    print(f'\n=== Sheet: {sheet_name} ===')
    print(f'Rows: {sheet.max_row}, Columns: {sheet.max_column}')
    
    # Print first 30 rows to understand structure
    for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        print(f'Row {i}: {row}')
